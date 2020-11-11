import openpyxl

def accountVerification(q, _email, _password):
    for i in q:
        if i[2] == _email and i[3] == _password:
            print("Welcome, " + i[0] + " " + i[1])
            return True
    return False

def homeScreen():
    s = input("What would you like to access? (Type tasks to access your tasks, expenses to track your expenses, or '0' to exit the application): ")
    return s

def tasksHomeScreen():
    s = input("To check all your tasks enter '1', to add a task enter '2', to mark a task for completion enter '3', and finally to go back to the home screen enter '4': ")
    return s

def tasks1(filename):
    file = open(filename, 'r')
    tasks = file.readlines()
    i = 1
    for line in tasks:
        print(str(i) + ". " + line + '\n')
        i += 1
    file.close()

def tasks2(filename):
    file = open(filename, 'a')
    x = True
    while x:
        task = input("Enter the task you want to add (type 0 if you are done entering your tasks): ")
        if task == '0':
            x = False
        else:
            file.write(task + "\n")

    file.close()

def tasks3(filename):
    tasks1(filename)
    file = open(filename, 'r')
    lines = file.readlines()
    file.close()
    x = True
    while x:
        completed = input("Enter the number of the task that you have completed (type 0 if you are done): ")
        if completed != '0':
            del lines[int(completed)-1]
            file1 = open(filename, 'w+')
            for line in lines:
                file1.write(line)
            file1.close()
            tasks1(filename)
        else:
            x = False

def expensesHomeScreen():
    s = input("To check all your expenses in the past 30 days type '1', to add today's expenses type '2', and finally to go back to the home screen type '3': ")
    return s

def expenses1(sheetName):
    filepath = "C:/Users/Mohammad/Desktop/LoginSystem/Expenses.xlsx"
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.get_sheet_by_name(sheetName)
    max_row = sheet.max_row
    max_column = sheet.max_column
    for i in range (2, max_column+1):
        cell = sheet.cell(row=1, column=i)
        print(str(cell.value), end= ": ")
        for j in range(2, max_row+1):
            cell = sheet.cell(row = j, column = i)
            if cell.value == None:
                continue
            else:
                print(str(cell.value), end=", ")

        print('\n')
    wb.save("Expenses.xlsx")

def expenses2(sheetName):
    filepath = "C:/Users/Mohammad/Desktop/LoginSystem/Expenses.xlsx"
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.get_sheet_by_name(sheetName)
    date = input("Enter the date of the day you would like to add expenses to (mm/dd/yyyy): ")
    rDate = 1
    rMax = sheet.max_row
    i = rMax - 2
    d = sheet.max_column
    sheet.cell(row = 1, column = 1).value = 0
    x = True
    while x:
        expense = input("Enter the expense you want to add (type 0 if you are done entering your expenses): ")
        if expense == '0':
            x = False
        else:
            sheet.cell(row = rMax - i, column = d+1).value = (expense)
            i -= 1
            wb.save("Expenses.xlsx")

    sheet.cell(row=rDate, column=d+1).value = date
    wb.save("Expenses.xlsx")



